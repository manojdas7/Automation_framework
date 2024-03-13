import openpyxl
from openpyxl.styles import PatternFill


def GetRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def GetColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnnum).value


def WriteData(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnnum).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',fill_type= 'solid')
    sheet.cell(rownum,columnnum).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='ff0000',
                            end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, columnnum).fill = greenFill
    workbook.save(file)















